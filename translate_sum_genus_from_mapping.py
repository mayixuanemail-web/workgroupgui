"""
属名翻译脚本（拉丁属名 -> 中文属名）

功能：
- 从映射表（XLSX）读取“拉丁属名 -> 中文属名”对照
- 扫描当前文件夹中的 Excel 文件
- 在原 Excel 中新增“中文属名”列

依赖：openpyxl
安装：pip install openpyxl
"""

from pathlib import Path
from openpyxl import load_workbook


# ======== 配置区域（按需修改）========
# 对照表所在目录（相对当前工作目录）
MAPPING_BASE_DIR = Path("mapping")

# 若文件名包含某关键词，则使用对应对照表
MAPPING_BY_KEYWORD = {
    "Archaea": "Archaea.xlsx",
    "Bacteria": "Bacteria.xlsx",
    "Fungi": "Fungi.xlsx",
    "Invertebrate": "Invertebrate.xlsx",
    "Mitochondrion": "Mitochondrion.xlsx",
    "Plant": "Plant.xlsx",
    "Plastid": "Plastid.xlsx",
    "Protozoa": "Protozoa.xlsx",
    "Vertebrate_Mammalian": "Vertebrate_Mammalian.xlsx",
    "Vertebrate_Other": "Vertebrate_Other.xlsx",
    "Viral": "Viral.xlsx",
}

DEFAULT_MAPPING_FILE = MAPPING_BASE_DIR / "genus_mapping.xlsx"  # 默认属名对照表文件
MAPPING_SHEET = None  # None 表示默认第一张表
MAPPING_LATIN_HEADER = "属"  # 对照表中拉丁属名列标题（找不到则用A列）
MAPPING_CN_HEADER = "中文属名"  # 对照表中中文属名列标题（找不到则用B列）

TARGET_KEYWORDS = []  # 仅处理包含这些关键词的文件名；留空=处理所有
GENUS_COL_HEADER = "属"  # 待翻译 Excel 中“拉丁属名”列标题
CN_COL_HEADER = "中文属名"  # 新增的中文列标题

SKIP_HEADERS = True  # 是否跳过表头行
HEADER_ROW_INDEX = 1
# =====================================


def _find_column_by_header(ws, header_name: str) -> int | None:
    """在表头行中查找列索引（1-based）。找不到返回 None。"""
    if not header_name:
        return None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=HEADER_ROW_INDEX, column=col).value == header_name:
            return col
    return None


def load_mapping(mapping_path: Path) -> dict:
    """加载属名对照表为字典：latin -> chinese"""
    wb = load_workbook(mapping_path, data_only=True)
    ws = wb.active if MAPPING_SHEET is None else wb[MAPPING_SHEET]

    latin_col = _find_column_by_header(ws, MAPPING_LATIN_HEADER) or 1
    cn_col = _find_column_by_header(ws, MAPPING_CN_HEADER) or 2

    mapping = {}
    start_row = HEADER_ROW_INDEX + 1 if SKIP_HEADERS else 1
    for row in range(start_row, ws.max_row + 1):
        latin = ws.cell(row=row, column=latin_col).value
        chinese = ws.cell(row=row, column=cn_col).value
        if latin:
            mapping[str(latin).strip()] = str(chinese).strip() if chinese is not None else ""

    wb.close()
    return mapping


def _get_mapping_for_file(excel_path: Path, mapping_cache: dict) -> tuple[dict, Path] | tuple[None, None]:
    """根据文件名关键词选择对照表，并从缓存中获取映射。"""
    selected_mapping_file = None
    for keyword, mapping_file in MAPPING_BY_KEYWORD.items():
        if keyword in excel_path.name:
            selected_mapping_file = MAPPING_BASE_DIR / Path(mapping_file)
            break

    if selected_mapping_file is None:
        selected_mapping_file = DEFAULT_MAPPING_FILE

    if not selected_mapping_file.exists():
        print(f"❌ 找不到对照表文件: {selected_mapping_file}")
        return None, None

    if selected_mapping_file not in mapping_cache:
        mapping_cache[selected_mapping_file] = load_mapping(selected_mapping_file)

    return mapping_cache[selected_mapping_file], selected_mapping_file


def translate_excel_file(excel_path: Path, mapping: dict):
    """在 Excel 中新增中文属名列"""
    wb = load_workbook(excel_path)
    ws = wb.active

    genus_col = _find_column_by_header(ws, GENUS_COL_HEADER)
    if genus_col is None:
        print(f"⚠️ 未找到列标题 '{GENUS_COL_HEADER}'，跳过: {excel_path.name}")
        wb.close()
        return

    # 找到/创建中文列
    cn_col = _find_column_by_header(ws, CN_COL_HEADER)
    if cn_col is None:
        cn_col = ws.max_column + 1
        ws.cell(row=HEADER_ROW_INDEX, column=cn_col, value=CN_COL_HEADER)

    start_row = HEADER_ROW_INDEX + 1 if SKIP_HEADERS else 1
    updated = 0

    for row in range(start_row, ws.max_row + 1):
        latin = ws.cell(row=row, column=genus_col).value
        if not latin:
            continue
        latin_key = str(latin).strip()
        chinese = mapping.get(latin_key, "")
        ws.cell(row=row, column=cn_col, value=chinese)
        updated += 1

    wb.save(excel_path)
    wb.close()
    print(f"✅ 已处理: {excel_path.name}，写入 {updated} 行中文属名")


def main():
    current_dir = Path("files_debug")
    for number_dir in sorted(current_dir.iterdir()):
        if not number_dir.is_dir():
            continue

        print(f"📂 当前目录: {number_dir}")

        mapping_cache: dict[Path, dict] = {}

        mapping_files_to_skip = {DEFAULT_MAPPING_FILE.name, *[Path(v).name for v in MAPPING_BY_KEYWORD.values()]}
        excel_files = [p for p in number_dir.glob("*.xlsx") if p.name not in mapping_files_to_skip]
        if TARGET_KEYWORDS:
            excel_files = [p for p in excel_files if any(k in p.name for k in TARGET_KEYWORDS)]

        if not excel_files:
            print("⚠️ 未找到需要处理的 Excel 文件")
            return

        for excel_file in sorted(excel_files):
            mapping, mapping_file = _get_mapping_for_file(excel_file, mapping_cache)
            if not mapping:
                continue
            print(f"📘 使用对照表: {mapping_file}")
            translate_excel_file(excel_file, mapping)


if __name__ == "__main__":
    main()
