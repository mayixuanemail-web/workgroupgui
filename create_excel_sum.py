"""生成汇总xlsx文件，包含所有part目录的分类结果

提供函数 `append_xlsx_to_summary`：
- 将一个源 `.xlsx` 的数据（带样式）追加到汇总文件中
- 保留单元格填充颜色（以及常见样式）
- 源文件与汇总文件有相同表头，只保留一个表头
- 如果指定列的单元格为红色（ff0000），则跳过该行

使用示例：
    from create_excel_sum import append_xlsx_to_summary
    append_xlsx_to_summary('src.xlsx', 'summary.xlsx', check_col='C')
"""

import os
from copy import copy
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils import column_index_from_string

def append_xlsx_to_summary(src_path, summary_path, check_col, header_rows=1, sheet_name=None, red_hex='FF0000'):
    """
    将 `src_path` 的内容追加到 `summary_path`的 "summary_name"（若不存在则创建）。
    参数：
    - src_path: 源 xlsx 文件路径
    - summary_name: 汇总文件名称
    - summary_path: 汇总 xlsx 文件路径
    - check_col: 要检查红色的列，可以是列字母（如 'C'）或 1-based 列索引（如 3）
    - header_rows: 表头行数（默认 1），只保留一次表头
    - sheet_name: 指定的 sheet 名称（默认使用第一个 sheet）
    - red_hex: 红色的十六进制值（不区分大小写，默认 'FF0000'）
    行为：
    - 复制每个单元格的值及常见样式（填充、字体、边框、对齐、数字格式、保护）
    - 若 check_col 对应单元格的填充色为 red_hex（支持 ARGB 格式如 '00FF0000' 或 'FF0000'），则跳过该行
    """

    if isinstance(check_col, str):
        check_col_idx = column_index_from_string(check_col)
    else:
        check_col_idx = int(check_col)

    red_hex = red_hex.strip().upper()

    src_wb = load_workbook(src_path)
    src_ws = src_wb[sheet_name] if sheet_name and sheet_name in src_wb.sheetnames else src_wb.active

    if os.path.exists(summary_path):
        sum_wb = load_workbook(summary_path)
        sum_ws = sum_wb[sheet_name] if sheet_name and sheet_name in sum_wb.sheetnames else sum_wb.active
        summary_exists = True
    else:
        sum_wb = Workbook()
        sum_ws = sum_wb.active
        if sheet_name:
            sum_ws.title = sheet_name
        summary_exists = False

    # 读取源表头（只比较第一行以判断是否已存在表头）
    def row_values(ws, row_idx):
        return ["" if c.value is None else str(c.value) for c in ws[row_idx]]

    src_header = row_values(src_ws, 1)
    sum_header = row_values(sum_ws, 1) if sum_ws.max_row >= 1 else []

    dest_row = sum_ws.max_row + 1 if summary_exists and sum_ws.max_row > 0 else 1

    # 如果汇总文件不存在或表头不同，则复制表头（header_rows 行）
    if not summary_exists or sum_header != src_header:
        for r in range(1, header_rows + 1):
            src_row = list(src_ws[r])
            for c_idx, src_cell in enumerate(src_row, start=1):
                tgt = sum_ws.cell(row=dest_row, column=c_idx, value=src_cell.value)
                try:
                    tgt.font = copy(src_cell.font)
                    tgt.border = copy(src_cell.border)
                    tgt.fill = copy(src_cell.fill)
                    tgt.number_format = src_cell.number_format
                    tgt.protection = copy(src_cell.protection)
                    tgt.alignment = copy(src_cell.alignment)
                except Exception:
                    # 忽略不能复制的样式
                    pass
            dest_row += 1

    # 判断单元格是否为指定红色（支持 ARGB 或 RGB 表示）
    def cell_is_red(cell):
        try:
            fg = getattr(cell.fill, 'fgColor', None)
            if fg is None:
                return False
            rgb = getattr(fg, 'rgb', None)
            if not rgb:
                return False
            rgb = rgb.upper()
            # 支持 'FF0000' 或 '00FF0000' 等格式，检查结尾
            return rgb.endswith(red_hex)
        except Exception:
            return False

    # 复制数据行（跳过 header_rows）
    for row in src_ws.iter_rows(min_row=header_rows + 1, max_row=src_ws.max_row):
        check_cell = row[check_col_idx - 1]
        if cell_is_red(check_cell):
            continue

        for c_idx, src_cell in enumerate(row, start=1):
            tgt = sum_ws.cell(row=dest_row, column=c_idx, value=src_cell.value)
            try:
                tgt.font = copy(src_cell.font)
                tgt.border = copy(src_cell.border)
                tgt.fill = copy(src_cell.fill)
                tgt.number_format = src_cell.number_format
                tgt.protection = copy(src_cell.protection)
                tgt.alignment = copy(src_cell.alignment)
            except Exception:
                pass
        dest_row += 1

    sum_wb.save(summary_path)
def batch_append_to_summary(base_path, check_col):
    """
    批量将指定目录下所有Excel文件的内容追加到汇总文件中
    base_path: 基础路径
    check_col: 要检查红色的列，可以是列字母（如 'C'）或 1-based 列索引（如 3）
    summary_name: 汇总文件名称（默认 'summary.xlsx'）
    """
    base_path = Path(base_path)
    # 遍历所有类别目录 (xxxx)
    for number_dir in sorted(base_path.iterdir()):
        if not number_dir.is_dir():
            continue
        for category_dir in sorted(number_dir.iterdir()):
            if not category_dir.is_dir():
                continue
            # 构建汇总文件路径
            summary_name = number_dir.name + '_' + category_dir.name + "_summary.xlsx"
            summary_path = number_dir / summary_name
            # 遍历所有 partxx 目录
            for part_dir in sorted(category_dir.iterdir()):
                if not part_dir.is_dir() or not part_dir.name.startswith("part"):
                    continue
                # 构建 species_taxonomy_table 目录路径
                table_dir = part_dir / "species_taxonomy_table"
                if not table_dir.exists():
                    continue
                # 查找所有xlsx文件
                xlsx_files = list(table_dir.glob("*.xlsx"))
                for xlsx_file in xlsx_files:
                    print(f"\n{'=' * 60}")
                    print(f"{category_dir.name}/{part_dir.name}")
                    try:
                        append_xlsx_to_summary(xlsx_file, summary_path, check_col)
                        print(f"  Appended data from '{xlsx_file}' to summary.")
                    except Exception as e:
                        print(f"  Error processing file: {e}")
def delete_xlsx_file(target_dir):
    """
    删除指定目录下所有的.xlsx格式文件
    :param target_dir: 目标目录的路径（可以是绝对路径或相对路径）
    :return: 无返回值，控制台输出删除结果
    """
    # 第一步：校验目标目录是否存在
    if not os.path.exists(target_dir):
        print(f"错误：目录 '{target_dir}' 不存在！")
        return
    
    # 第二步：校验目标路径是否是目录（避免传入的是文件路径）
    if not os.path.isdir(target_dir):
        print(f"错误：'{target_dir}' 不是一个有效的目录！")
        return
    
    # 第三步：遍历目录下的所有文件，筛选并删除.xlsx文件
    deleted_count = 0  # 统计删除的文件数量
    for file_name in os.listdir(target_dir):
        # 拼接文件的完整路径
        file_path = os.path.join(target_dir, file_name)
        
        # 只处理文件（排除子目录），且后缀是.xlsx
        if os.path.isfile(file_path) and file_name.lower().endswith('.xlsx'):
            try:
                # 删除文件
                os.remove(file_path)
                deleted_count += 1
                print(f"成功删除文件：{file_path}")
            except PermissionError:
                print(f"权限不足，无法删除文件：{file_path}")
            except FileNotFoundError:
                print(f"文件不存在，跳过：{file_path}")
            except Exception as e:
                print(f"删除文件 {file_path} 失败：{str(e)}")
    
    # 输出最终结果
    print(f"\n删除完成！本次共尝试删除 {deleted_count} 个.xlsx文件")

if __name__ == "__main__":
    print("=" * 60)
    # 批量处理模式
    base_path = "files_debug"
    check_col = "F"  # 检查红色的列，可以是列字母或1-based索引
    #summary_name = "summary.xlsx"  # 可自定义
    for number_dir in sorted(Path(base_path).iterdir()):
        delete_xlsx_file(number_dir)
    batch_append_to_summary(base_path, check_col)
    print("\n" + "=" * 60)