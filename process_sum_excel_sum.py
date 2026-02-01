"""
Excel处理脚本
功能：
1. 读取表头为属的列的数据，将相同字符串对应的reads列值求和
2. 保留相同字符串的第一行，reads列值改为原数据的求和
3. 覆盖原文件
"""
import os
import sys
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
def process_excel(file_path):
    '''
    处理xlsx文件
    1. 读取表头为属的列的数据，将相同字符串对应的reads列值求和
    2. 保留相同字符串的第一行，reads列值改为原数据的求和
    3.如果属列的原单元格有颜色则保留颜色信息
    '''
    import openpyxl
    from openpyxl.styles import PatternFill

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # 获取表头
        headers = [cell.value for cell in ws[1]]
        print(f"  表头: {headers}")
        
        if '属' not in headers or 'reads' not in headers:
            print(f"  ❌ 未找到'属'或'reads'列: {file_path}")
            return False

        genus_col_idx = headers.index('属') + 1  # openpyxl是1-based
        reads_col_idx = headers.index('reads') + 1
        print(f"  属列索引: {genus_col_idx}, reads列索引: {reads_col_idx}")

        # 读取所有数据，记录颜色
        data = []
        color_map = {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            genus = row[genus_col_idx-1].value
            reads = row[reads_col_idx-1].value
            # 记录颜色
            fill = row[genus_col_idx-1].fill
            color = fill.fgColor.rgb if fill and fill.fgColor.type == 'rgb' else None
            color_map.setdefault(genus, color)
            data.append({
                'row': row,
                'genus': genus,
                'reads': reads,
                'color': color
            })
        
        print(f"  读取数据行数: {len(data)}")

        # 按属分组求和
        from collections import OrderedDict
        genus_sum = OrderedDict()
        for d in data:
            genus = d['genus']
            reads = d['reads']
            # 跳过空的属名
            if genus is None or str(genus).strip() == '':
                continue
            try:
                reads = float(reads) if reads is not None else 0
            except (TypeError, ValueError):
                reads = 0
            if genus not in genus_sum:
                genus_sum[genus] = {'sum': reads, 'first_row': d['row'], 'color': d['color']}
            else:
                genus_sum[genus]['sum'] += reads

        print(f"  分组数量: {len(genus_sum)}")

        # 清空原数据行
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # 写入新数据，保留首行和颜色
        write_count = 0
        for genus, info in genus_sum.items():
            row = [cell.value for cell in info['first_row']]
            row[genus_col_idx-1] = genus
            row[reads_col_idx-1] = info['sum']
            ws.append(row)
            write_count += 1
            # 设置属列颜色
            new_row_idx = ws.max_row
            cell = ws.cell(row=new_row_idx, column=genus_col_idx)
            if info['color']:
                cell.fill = PatternFill(fill_type='solid', fgColor=info['color'])

        wb.save(file_path)
        print(f"  ✅ 处理成功！分组: {write_count}, 文件: {file_path}")
        return True
        
    except Exception as e:
        print(f"  ❌ 处理异常: {file_path}")
        print(f"     错误信息: {type(e).__name__}: {str(e)}")
        import traceback
        traceback.print_exc()
        return False
def batch_process_excel_in_directory(base_path, success=0, fail=0):
    """
    批量处理指定目录下所有Excel文件
    base_path: 基础路径
    """
    # 遍历所有类别目录 (xxxx)
    for number_dir in sorted(base_path.iterdir()):
        if not number_dir.is_dir():
            continue
        # 查找所有xlsx文件
        xlsx_files = list(number_dir.glob("*.xlsx")) 
        for xlsx_file in xlsx_files:
            print(f"\n{'=' * 60}")
            print(f"处理文件: {xlsx_file.name}")
            try:
                if process_excel(xlsx_file):
                    success += 1
                else:
                    fail += 1
            except Exception as e:
                print(f"  ❌ 未捕获的异常 {xlsx_file}: {e}")
                import traceback
                traceback.print_exc()
                fail += 1
    return success, fail
if __name__ == "__main__":
    print("=" * 60)
    # 批量处理模式
    base_path = "files_debug"
    #base_path = r"C:\Users\ma\Desktop\workgroup2\files_debug\1_3_1_3_fastp" 
    success, fail = batch_process_excel_in_directory(Path(base_path))          
    print("\n" + "=" * 60)
    print(f"Batch processing completed. Success: {success}, Failed: {fail}")
