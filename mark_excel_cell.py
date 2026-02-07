#读取某个目录下的xlsx文件，读取其中xxx表头列下的文本
#读取另外一个目录下的XLSX文件，标记对应单元格颜色为x色
import pandas as pd
import openpyxl
import os
from pathlib import Path
from openpyxl.styles import PatternFill
import typing
def mark_excel_cell(file_path, target_col, reference_data, fill_color):
    """
    标记Excel文件中指定列的单元格颜色
    file_path: Excel文件路径
    target_col: 需要标记颜色的列名
    reference_data: 参考数据列表，包含需要标记的单元格值
    fill_color: 填充颜色（十六进制字符串，如'FFFF00'表示黄色）
    """
    # 1. 读取 Excel 文件
    df = pd.read_excel(file_path)
    
    # 检查目标列是否存在
    if target_col not in df.columns:
        raise ValueError(f"列名 {target_col} 在 Excel 文件中不存在！")
    
    # 2. 用 openpyxl 加载工作簿，准备修改样式
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # 假设处理第一个工作表
    
    # 定义填充样式
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # 3. 获取目标列在 Excel 中的字母位置（例如 A, B, C...）
    col_idx = df.columns.get_loc(target_col) + 1  # 转成 openpyxl 的列索引（从1开始）
    excel_col = ws.cell(row=1, column=col_idx).column_letter  # 得到列字母（如 'D'）
    
    # 4. 遍历数据行，标记颜色
    # 表头行是第1行，数据从第2行开始
    for row_num in range(2, len(df) + 2):
        raw_value = ws[f"{excel_col}{row_num}"].value
        cell_value = _sanitize_text(raw_value)
        pros_cell_value = _sanitize_text(raw_value)
        if pros_cell_value and pros_cell_value in reference_data:
            ws[f"{excel_col}{row_num}"].fill = fill
    
    # 5. 保存修改后的文件
    wb.save(file_path)
    print(f"已完成标记，文件已保存到: {file_path}")

def get_reference_data_from_file(ref_file_path, ref_col):
    """
    从参考Excel文件中读取指定列的所有单元格值，作为参考数据列表
    ref_file_path: 参考Excel文件路径
    ref_col: 参考列名
    """
    df = pd.read_excel(ref_file_path)
    if ref_col not in df.columns:
        print(f"  Error: Reference column '{ref_col}' not found in the reference file.")
        return []
    raw_vals = df[ref_col].dropna().astype(str).tolist()
    cleaned = []
    for v in raw_vals:
        s = _sanitize_text(v)
        if s:
            cleaned.append(s)
    # 去重并保持顺序
    seen = set()
    uniq = []
    for item in cleaned:
        if item not in seen:
            seen.add(item)
            uniq.append(item)
    return uniq


def _sanitize_text(val: typing.Any) -> str:
    """
    移除字符串中的非字母字符，仅保留字母（支持 Unicode 字母），并转为小写。
    例如："abc-123" -> "abc"；"A_B" -> "ab"。
    """
    if val is None:
        return ""
    s = str(val)
    letters = [ch for ch in s if ch.isalpha()]
    return ("".join(letters)).lower()

def batch_mark_column_in_directory(base_path, target_col, ori_col,fill_color, success=0, fail=0):
    """
    批量标记指定目录下所有Excel文件中的某一列的单元格颜色
    base_path: 基础路径
    target_col: 需要标记颜色的列名
    reference_data: 参考数据列表，包含需要标记的单元格值
    fill_color: 填充颜色（十六进制字符串，如'FFFF00'表示黄色）
    """
    # 遍历所有类别目录 (xxxx)
    for number_dir in sorted(base_path.iterdir()):
        if not number_dir.is_dir():
            continue
        for category_dir in sorted(number_dir.iterdir()):
            if not category_dir.is_dir():
                continue
            # 遍历所有 partxx 目录
            for part_dir in sorted(category_dir.iterdir()):
                if not part_dir.is_dir() or not part_dir.name.startswith("part"):
                    continue
                
                # 构建 species_taxonomy_table 目录路径
                table_dir = part_dir / "species_taxonomy_table"

                # 查找 part_dir 下包含"分类结果"的 xlsx 文件
                xlsx_files_in_part = list(part_dir.glob("*分类结果.xlsx"))
                if not xlsx_files_in_part:
                    print(f"  Error: No '分类结果.xlsx' file found in '{part_dir}', skipping...")
                    continue
                
                # 使用第一个找到的文件
                xlsx_dir = xlsx_files_in_part[0]
                if len(xlsx_files_in_part) > 1:
                    print(f"  Warning: Multiple '分类结果.xlsx' files found in '{part_dir}', using: {xlsx_dir.name}")
                
                if not table_dir.exists():
                    print(f"  Error: Table directory '{table_dir}' not found, skipping...")
                    continue
                # 读取参考数据
                reference_data = get_reference_data_from_file(xlsx_dir, ori_col)
                # 查找所有xlsx文件
                xlsx_files = list(table_dir.glob("*.xlsx")) 
                for xlsx_file in xlsx_files:
                    print(f"\n{'=' * 60}")
                    print(f"{category_dir.name}/{part_dir.name}")
                    try:
                        mark_excel_cell(xlsx_file, target_col, reference_data, fill_color)
                        success += 1
                    except Exception as e:
                        print(f"  Error processing file {xlsx_file}: {e}")
                        fail += 1
    return success, fail
if __name__ == "__main__":
    
    for tran in range(1, 4):
        if tran==1:
            print("=" * 60)
            # 批量处理模式
            base_path = "files_debug"
            target_col = "属"
            ori_col = "好"
            fill_color = "FFFF00"  # 黄色
            success, fail = batch_mark_column_in_directory(Path(base_path), target_col, ori_col, fill_color)          
            print("\n" + "=" * 60)
            print(f"Batch processing completed. Success: {success}, Failed: {fail}")
            tran+=1
        if tran==2:
            print("=" * 60)
            # 批量处理模式
            base_path = "files_debug"
            target_col = "属"
            ori_col = "一到四个异常点"
            fill_color = "00FF00"  # 绿色
            success, fail = batch_mark_column_in_directory(Path(base_path), target_col, ori_col, fill_color)          
            print("\n" + "=" * 60)
            print(f"Batch processing completed. Success: {success}, Failed: {fail}")
            tran+=1
        if tran==3:
            print("=" * 60)
            # 批量处理模式
            base_path = "files_debug"
            target_col = "属"
            ori_col = "平"
            fill_color = "FF0000"  # 红色
            success, fail = batch_mark_column_in_directory(Path(base_path), target_col, ori_col, fill_color)          
            print("\n" + "=" * 60)
            print(f"Batch processing completed. Success: {success}, Failed: {fail}")
            tran+=1
            break