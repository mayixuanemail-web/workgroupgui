"""
Excel F、G 列数据提取与汇总工具

功能说明：
1. 扫描当前文件夹中所有文件名包含 "xxx" 的 Excel 文件
2. 提取每个文件的 F 列和 G 列数据（自动跳过表头）
3. 合并所有数据到新文件 xxx.xlsx
4. 对 F 列去重，保留首次出现的行

依赖：openpyxl
安装：pip install openpyxl
"""

from pathlib import Path
from openpyxl import Workbook, load_workbook


def extract_fg_columns(excel_path: Path) -> list:
    """
    从 Excel 文件中提取 F、H 列数据
    
    Args:
        excel_path: Excel 文件路径
    
    Returns:
        list: [(F值, H值), ...] 数据列表
    """
    try:
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        
        data = []
        # 从第2行开始读取（跳过表头）
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=8, values_only=True):
            f_val, g_val, h_val = row
            # 只保留 F 列和 H 列都有内容的行
            if f_val is not None and f_val != "" and h_val is not None and h_val != "":
                data.append((f_val, h_val))
        
        wb.close()
        return data
    
    except Exception as e:
        print(f"  ❌ 读取失败: {e}")
        return []


def merge_and_deduplicate(all_data: list) -> list:
    """
    对数据按 F 列去重，H 列保留较短的值
    
    Args:
        all_data: 所有数据列表
    
    Returns:
        list: 去重后的数据
    """
    f_dict = {}  # 用字典存储 F值 -> (H值, 长度)
    
    for f_val, h_val in all_data:
        h_str = str(h_val) if h_val is not None else ""
        h_len = len(h_str)
        
        if f_val not in f_dict:
            f_dict[f_val] = (h_val, h_len)
        else:
            # 比较长度，保留较短的
            current_h, current_len = f_dict[f_val]
            if h_len < current_len:
                f_dict[f_val] = (h_val, h_len)
    
    # 转换回列表形式
    unique_data = [(f_val, h_val) for f_val, (h_val, _) in f_dict.items()]
    return unique_data


def create_output_excel(data: list, output_path: Path):
    """
    创建输出 Excel 文件
    
    Args:
        data: 数据列表
        output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    
    # 写入表头
    ws['A1'] = 'F列'
    ws['B1'] = 'H列'
    
    # 写入数据（从第2行开始）
    for idx, (f_val, h_val) in enumerate(data, start=2):
        ws.cell(row=idx, column=1, value=f_val)
        ws.cell(row=idx, column=2, value=h_val)
    
    wb.save(output_path)
    print(f"✅ 已保存汇总文件: {output_path}")
    print(f"   共 {len(data)} 行数据（已去重）")


def process_keyword(keyword: str, current_dir: Path):
    """
    处理单个关键词的所有文件
    
    Args:
        keyword: 文件名关键词
        current_dir: 当前目录
    """
    output_filename = f"{keyword}.xlsx"
    
    print(f"\n{'='*60}")
    print(f"🔍 处理关键词: {keyword}")
    print(f"{'='*60}\n")
    
    # 搜索所有包含关键词的 Excel 文件
    excel_files = []
    for pattern in ["*.xlsx", "*.xls"]:
        for file in current_dir.glob(pattern):
            if keyword in file.name and file.name != output_filename:
                excel_files.append(file)
    
    if not excel_files:
        print(f"⚠️ 未找到包含 '{keyword}' 的 Excel 文件\n")
        return
    
    print(f"找到 {len(excel_files)} 个匹配文件:\n")
    
    # 提取所有数据
    all_data = []
    for excel_file in sorted(excel_files):
        print(f"📄 处理: {excel_file.name}")
        data = extract_fg_columns(excel_file)
        if data:
            print(f"  ✅ 提取 {len(data)} 行数据")
            all_data.extend(data)
        else:
            print(f"  ⏭️ 无有效数据，跳过")
    
    if not all_data:
        print(f"\n⚠️ 关键词 '{keyword}' 的所有文件均无有效数据\n")
        return
    
    print(f"\n📊 合并前总数据: {len(all_data)} 行")
    
    # 去重
    unique_data = merge_and_deduplicate(all_data)
    print(f"📊 去重后数据: {len(unique_data)} 行\n")
    
    # 创建输出文件
    output_path = current_dir / output_filename
    create_output_excel(unique_data, output_path)


def main():
    """主函数"""
    # 🔧 配置参数：11个生物分类关键词
    KEYWORDS = [
        "Archaea",
        "Bacteria",
        "Fungi",
        "Invertebrate",
        "Mitochondrion",
        "Plant",
        "Plastid",
        "Protozoa",
        "Vertebrate_Mammalian",
        "Vertebrate_Other",
        "Viral",
    ]
    
    current_dir = Path("translate")
    print(f"📂 当前工作目录: {current_dir}")
    print(f"📝 将处理 {len(KEYWORDS)} 个关键词")
    
    # 逐个处理每个关键词
    for keyword in KEYWORDS:
        process_keyword(keyword, current_dir)
    
    print(f"\n{'='*60}")
    print("🎉 全部处理完成！")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
