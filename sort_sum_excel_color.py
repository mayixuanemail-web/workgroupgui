"""
按表头为x的列单元格颜色排序Excel
排序顺序: 无色>橙色>黄色>绿色>红色
重排后保存原颜色信息
"""
import pandas as pd
import openpyxl
import os
from pathlib import Path
from openpyxl.styles import PatternFill, Font, Alignment, Border
def sort_excel_color(file_path, target_col):
    """
    按指定列的单元格颜色排序Excel文件
    file_path: Excel文件路径
    target_col: 需要排序颜色的列名
    保存到原文件
    """
    print(f"Processing: {file_path}")
    
    # 1. 读取 Excel 文件（仅用于校验列和获取行数）
    df = pd.read_excel(file_path)
    
    # 检查目标列是否存在
    if target_col not in df.columns:
        raise ValueError(f"列名 {target_col} 在 Excel 文件中不存在！")
    
    print(f"  表头列数: {len(df.columns)}")
    print(f"  数据行数: {len(df)}")
    print(f"  排序列: {target_col}")
    
    # 2. 用 openpyxl 加载工作簿，读取填充色和值
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # 获取目标列的数字索引（从1开始）
    col_idx = df.columns.get_loc(target_col) + 1  
    print(f"  目标列索引: {col_idx}")
    
    # 3. 定义颜色优先级 + 收集每行数据（值+颜色）
    color_order = {
        '00000000': 0,      # 无色（黑色，默认）
        'FF000000': 0,      # 无色（带透明度）
        '00FFA500': 1,      # 橙
        'FFFFA500': 1,      # 橙（带透明度）
        '00FFFF00': 2,      # 黄
        'FFFFFF00': 2,      # 黄（带透明度）
        '0000FF00': 3,      # 绿
        'FF00FF00': 3,      # 绿（带透明度）
        '00FF0000': 4,      # 红
        'FFFF0000': 4,      # 红（带透明度）
    }
    
    row_data = []  # 存储：(颜色优先级, 整行值, 目标列填充色)
    
    # 遍历数据行（表头行1，数据从2开始）
    for row_num in range(2, len(df) + 2):
        # 获取目标列单元格的填充色
        target_cell = ws.cell(row=row_num, column=col_idx)
        fill = target_cell.fill
        
        # 安全读取颜色
        fg_color = '00000000'  # 默认无色
        if fill and fill.fgColor:
            try:
                if fill.fgColor.type == "rgb" and fill.fgColor.rgb:
                    fg_color = fill.fgColor.rgb
                elif fill.fgColor.type == "theme":
                    # 如果是主题颜色，尝试转换
                    fg_color = '00000000'
            except Exception as e:
                print(f"    读取颜色异常 (行{row_num}): {e}")
        
        # 获取颜色优先级
        color_priority = color_order.get(fg_color, 0)
        
        # 获取整行的单元格值
        row_values = []
        for col_num in range(1, ws.max_column + 1):
            row_values.append(ws.cell(row=row_num, column=col_num).value)
        
        # 获取属列的内容用于日志
        genus_value = row_values[col_idx - 1] if col_idx <= len(row_values) else "N/A"
        print(f"    行{row_num}: 属='{genus_value}' | 颜色={fg_color} | 优先级={color_priority}")
        
        # 存储关键信息：颜色优先级、整行值、目标列填充色
        row_data.append((color_priority, row_values, fg_color))
    
    print(f"  收集行数: {len(row_data)}")
    
    # 4. 按颜色优先级排序
    print(f"  排序前顺序: {[x[0] for x in row_data]}")
    row_data.sort(key=lambda x: x[0])
    print(f"  排序后顺序: {[x[0] for x in row_data]}")
    
    # 5. 清空原工作表数据（保留表头），重新写入排序后的数据
    # 清空数据行（从第2行开始）
    for row_num in range(2, ws.max_row + 1):
        for col_num in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col_num).value = None
            ws.cell(row=row_num, column=col_num).fill = PatternFill(fill_type=None)  # 清空填充色
    
    # 写入排序后的数据 + 恢复目标列填充色
    new_row_num = 2
    for _, row_values, fg_color in row_data:
        # 写入整行值
        for col_num in range(1, len(row_values) + 1):
            ws.cell(row=new_row_num, column=col_num).value = row_values[col_num - 1]
        
        # 恢复目标列的填充色（核心：只保留排序依据的颜色）
        if fg_color != '00000000':  # 非无色才填充
            target_cell = ws.cell(row=new_row_num, column=col_idx)
            target_cell.fill = PatternFill(
                start_color=fg_color,
                end_color=fg_color,
                fill_type='solid'
            )
        
        new_row_num += 1
    
    # 6. 保存并关闭工作簿
    wb.save(file_path)
    wb.close()
    
    print(f"  ✅ 文件排序完成并保存")
    return True

def batch_sort_color_in_directory(base_path, target_col, success=0, fail=0):
    """
    批量按指定目录下所有Excel文件中的某一列的单元格颜色排序
    base_path: 基础路径
    target_col: 需要排序颜色的列名
    """
    # 遍历所有类别目录 (xxxx)
    for number_dir in sorted(base_path.iterdir()):
        if not number_dir.is_dir():
            continue
        
        # 查找所有xlsx文件
        xlsx_files = list(number_dir.glob("*.xlsx")) 
        for xlsx_file in xlsx_files:
            print(f"\n{'=' * 60}")
            print(f"{number_dir.name}/{xlsx_file.name}")
            try:
                if sort_excel_color(xlsx_file, target_col):
                    success += 1
            except Exception as e:
                print(f"  ❌ 处理失败: {e}")
                import traceback
                traceback.print_exc()
                fail += 1
    return success, fail
if __name__ == "__main__":
    print("=" * 60)
    # 批量处理模式
    base_path = "files_debug"
    target_col = "属"
    success, fail = batch_sort_color_in_directory(Path(base_path), target_col)          
    print("\n" + "=" * 60)
    print(f"Batch processing completed. Success: {success}, Failed: {fail}")
    print("=" * 60)