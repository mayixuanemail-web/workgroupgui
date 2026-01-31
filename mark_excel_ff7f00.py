#从某个文件夹中提取pdf名字，并将xlsx涂色'#ff7f00'
import pandas as pd
import os
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
def extract_keyword_from_pdf_name(pdf_name):
    """
    从PDF文件名中提取属名关键字
    
    参数：
        pdf_name (str): PDF文件名，格式如 "1_5_1_43_fastp.Bacteria.part01.Acidovorax.pdf"
    
    返回：
        str 或 None: 提取的属名（如"Acidovorax"），提取失败则返回None
    
    文件名结构解析：
        1_5_1_43_fastp.Bacteria.part01.Acidovorax.pdf
        ├── 1_5_1_43_fastp  → 样本/项目标识
        ├── Bacteria        → 生物分类（细菌/真菌等）
        ├── part01          → 分区编号
        └── Acidovorax      → 属名（我们需要提取的部分）
    """
    # 第一步：移除.pdf后缀
    # "xxx.Acidovorax.pdf" → "xxx.Acidovorax"
    name = pdf_name.replace(".pdf", "")
    
    # 第二步：按点号分割字符串
    # "1_5_1_43_fastp.Bacteria.part01.Acidovorax" → ["1_5_1_43_fastp", "Bacteria", "part01", "Acidovorax"]
    parts = name.split(".")
    
    # 第三步：取最后一个部分（属名）
    # 要求至少有4个部分才认为是有效的文件名格式
    if len(parts) >= 4:
        return parts[-1]  # parts[-1] 表示列表的最后一个元素
    
    # 如果格式不符合预期，返回None
    return None

def mark_excel_cells(xlsx_path, keywords, col):
    """
    标记Excel文件中指定列的单元格：将黄色背景且内容匹配关键字的单元格改为橙色。
    
    参数：
        xlsx_path (str 或 Path): Excel文件路径
        keywords (list): 关键字列表
        col (str): 列名（如'A'、'B'等）
    """
    # 1. 打开Excel文件
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active  # 获取活动工作表
    
    # 黄色和橙色的十六进制表示
    YELLOW_COLOR = 'FFFF00'  # 黄色
    ORANGE_COLOR = 'FF7F00'  # 橙色
    
    # 将关键字列表转为集合以提高查找效率
    keyword_set = set(keywords) if isinstance(keywords, list) else {keywords}
    
    # 2. 遍历col列的每个单元格
    for row in range(1, ws.max_row + 1):
        cell = ws[f"{col}{row}"]
        
        # 3. 检查单元格内容是否等于关键字且这个单元格是黄色
        cell_value = str(cell.value).strip() if cell.value else ""
        
        # 检查单元格是否有填充色
        if cell.fill and cell.fill.start_color:
            cell_color = cell.fill.start_color.rgb
            
            # 检查是否为黄色（可能是'FFFF00'或'00FFFF00'格式）
            is_yellow = (cell_color == YELLOW_COLOR or 
                        cell_color == f"00{YELLOW_COLOR}" or
                        cell_color == f"FF{YELLOW_COLOR}")
            
            # 4. 如果匹配，将该单元格背景设为橙色
            if is_yellow and cell_value in keyword_set:
                from openpyxl.styles import PatternFill
                cell.fill = PatternFill(start_color=ORANGE_COLOR, 
                                       end_color=ORANGE_COLOR, 
                                       fill_type='solid')
    
    # 5. 保存并关闭文件
    wb.save(xlsx_path)
    wb.close()

def batch_mark_excel_cells(base_dir, excel_col, fill_color):
    """
    批量处理指定目录下的所有.xlsx文件，标记指定列的单元格颜色。
    
    参数：
        base_dir (str 或 Path): 包含.xlsx文件的目录路径
        excel_col (str): 需要标记颜色的列名
        fill_color (str): 填充颜色（十六进制字符串，如'FFFF00'表示黄色）
    """
    base_path = Path(base_dir)
    success = 0
    fail = 0
    
    for number_dir in base_path.iterdir():
        if not number_dir.is_dir():
            continue
        
        for category_dir in number_dir.iterdir():
            if not category_dir.is_dir():
                continue
            
            for part_dir in category_dir.iterdir():
                if not part_dir.is_dir():
                    continue
                
                # 构建 xlsx 文件路径
                xlsx_files = part_dir / "species_taxonomy_table"
                xlsx_files = list(xlsx_files.glob("*.xlsx"))
                if not xlsx_files:
                    print(f"  No .xlsx files found in {part_dir}, skipping...")
                    continue
                
                xlsx_file = xlsx_files[0]  # 假设每个 part_dir 只有一个 xlsx 文件
                
                try:
                    print(f"Processing: {xlsx_file}")
                    
                    # 提取该 part_dir 下所有 PDF 文件的属名关键字
                    pdf_dir= part_dir / "非常好"
                    pdf_files = list(pdf_dir.glob("*.pdf"))
                    keywords = []
                    for pdf_file in pdf_files:
                        keyword = extract_keyword_from_pdf_name(pdf_file.name)
                        if keyword:
                            keywords.append(keyword)
                    
                    # 标记 Excel 文件中的单元格
                    mark_excel_cells(xlsx_file, keywords, excel_col)
                    
                    success += 1
                except Exception as e:
                    print(f"  Error processing file {xlsx_file}: {e}")
                    fail += 1
    return success, fail
if __name__ == "__main__":
    print("=" * 60)
    # 批量处理模式
    base_dir = "files_debug" 
    excel_col = "F"  # 需要标记颜色的列名
    fill_color = "FF7F00"  # 橙色
    
    success, fail = batch_mark_excel_cells(base_dir, excel_col, fill_color)          
    print("\n" + "=" * 60)