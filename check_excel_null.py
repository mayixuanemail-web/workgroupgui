#检查每个excel中某个列是否有无色单元格
#将检查结果输出到C:\Users\ma\Desktop\workgroup2\result.txt
import pandas as pd
import openpyxl
import os
from pathlib import Path
def check_excel_null(file_path, target_col):
    """
    检查Excel文件中指定列是否有无色单元格
    file_path: Excel文件路径
    target_col: 需要检查的列名
    """
    print(f"Processing: {file_path}")
    
    # 1. 读取 Excel 文件
    df = pd.read_excel(file_path)
    
    # 检查目标列是否存在
    if target_col not in df.columns:
        raise ValueError(f"列名 {target_col} 在 Excel 文件中不存在！")
    
    # 2. 用 openpyxl 加载工作簿，准备检查样式
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # 假设处理第一个工作表
    
    # 3. 获取目标列在 Excel 中的字母位置（例如 A, B, C...）
    col_idx = df.columns.get_loc(target_col) + 1  # 转成 openpyxl 的列索引（从1开始）
    excel_col = ws.cell(row=1, column=col_idx).column_letter  # 得到列字母（如 'D'）
    
    # 4. 遍历数据行，检查无色单元格
    # 只有当单元格不是红/黄/绿时视为“无色”（包括无填充）
    null_count = 0
    allowed = {"FF0000", "FFFF00", "00FF00"}  # 允许的颜色（红、黄、绿），6位RGB大写
    def _extract_hex(val):
        if not val:
            return None
        s = str(val).upper()
        s = s.lstrip('#')
        if len(s) == 8 and s.startswith('FF'):
            s = s[2:]
        if len(s) > 6:
            s = s[-6:]
        return s if len(s) == 6 else None

    for row_num in range(2, len(df) + 2):
        cell = ws[f"{excel_col}{row_num}"]
        fill = cell.fill
        pattern = getattr(fill, 'patternType', None) or getattr(fill, 'fill_type', None)
        if not pattern:
            null_count += 1
            continue
        start_color = getattr(fill, 'start_color', None)
        rgb = None
        if start_color is not None:
            rgb = getattr(start_color, 'rgb', None) or getattr(start_color, 'index', None)
        hex6 = _extract_hex(rgb)
        # 如果提取不到有效颜色，或颜色不是红/黄/绿，则视为无色
        if hex6 not in allowed:
            null_count += 1
    
    print(f"  Found {null_count} null (uncolored) cells in column '{target_col}'")
    return null_count

def batch_check_null_in_directory(base_path, target_col, results_file):
    """
    批量检查指定目录下所有Excel文件中的某一列是否有无色单元格
    base_path: 基础路径
    target_col: 需要检查的列名
    results_file: 结果输出文件路径
    """
    with open(results_file, 'w', encoding='utf-8') as rf:
        # 遍历所有类别目录 (xxxx)
        for number_dir in sorted(base_path.iterdir()):
            if not number_dir.is_dir():
                continue
            for category_dir in sorted(number_dir.iterdir()):
                if not category_dir.is_dir():
                    continue
                # 遍历所有 partxx 目录
                for part_dir in sorted(category_dir.iterdir()):
                    if not part_dir.is_dir() or not part_dir.name.startswith("part"):
                        continue
                    
                    # 构建 species_taxonomy_table 目录路径
                    table_dir = part_dir / "species_taxonomy_table"
                    if not table_dir.exists():
                        continue
                    # 查找所有xlsx文件
                    xlsx_files = list(table_dir.glob("*.xlsx")) 
                    for xlsx_file in xlsx_files:
                        print(f"\n{'=' * 60}")
                        print(f"{category_dir.name}/{part_dir.name}")
                        try:
                            null_count = check_excel_null(xlsx_file, target_col)
                            if null_count > 0:
                                rf.write(f"{category_dir.name}/{part_dir.name}/{xlsx_file.name}: {null_count} null cells in column '{target_col}'\n")
                        except Exception as e:
                            print(f"  Failed to process {xlsx_file}: {e}")
                            rf.write(f"{category_dir.name}/{part_dir.name}/{xlsx_file.name}: Error - {e}\n")
    print(f"\nResults written to: {results_file}")
if __name__ == "__main__":
    print("=" * 60)
    # 批量处理模式
    base_path = "files_debug"
    target_col = "属"
    results_file = r"C:\Users\ma\Desktop\workgroup2\result.txt"
    batch_check_null_in_directory(Path(base_path), target_col, results_file)